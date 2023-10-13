"""
Perseo - Dispersiones Bancos
"""
import rich
import typer
from openpyxl import Workbook

from config.settings import get_settings
from lib.exceptions import MyAnyError
from perseo.dispersiones_bancos.loaders import leer
from perseo.dispersiones_bancos.searchers import buscar_rfc

app = typer.Typer()


@app.command()
def mostrar():
    """Mostrar dispersiones bancarias"""
    rich.print("Mostrar dispersiones bancarias")


@app.command()
def buscar(rfc: str):
    """Buscar un RFC"""
    rich.print("Buscar un RFC")

    # Obtener configuracion
    settings = get_settings()

    # Buscar RFC
    try:
        dispersiones = buscar_rfc(settings=settings, rfc=rfc)
    except MyAnyError as error:
        typer.secho(str(error), fg=typer.colors.RED)
        raise typer.Exit()

    # Bucle para mostrar las dispersiones
    for dispersion in dispersiones:
        # Mostrar resultado
        rich.print(f"Centro de trabajo: [green]{dispersion.persona.centro_trabajo_clave}[/green]")
        rich.print(f"RFC:               [green]{dispersion.persona.rfc}[/green]")
        rich.print(f"Nombre:            [green]{dispersion.persona.nombre}[/green]")
        rich.print(f"Municipio:         [green]{dispersion.persona.municipio.nombre}[/green]")
        rich.print(f"Plaza:             [green]{dispersion.persona.plaza}[/green]")
        rich.print(f"Sexo:              [white]{dispersion.persona.sexo}[/white]")
        rich.print()

        # Mostrar percepciones
        for percepcion in dispersion.percepciones_deducciones:
            if percepcion.concepto.p_d.value == "P":
                rich.print(f"{percepcion.concepto.descripcion}: [green]{percepcion.importe}[/green]")
        rich.print()

        # Mostrar deducciones
        for percepcion in dispersion.percepciones_deducciones:
            if percepcion.concepto.p_d.value == "D":
                rich.print(f"{percepcion.concepto.descripcion}: [red]{percepcion.importe}[/red]")
        rich.print()

        # Mostrar cantidades finales
        rich.print(f"Percepcion:        [blue]{dispersion.percepcion}[/blue]")
        rich.print(f"Deduccion:         [red]{dispersion.deduccion}[/red]")
        rich.print(f"Importe:           [green]{dispersion.importe}[/green]")
        rich.print(f"No. Cheque:        [gray]{dispersion.num_cheque}[/gray]")
        rich.print()


@app.command()
def guardar(rfc: str):
    """Crear un archivo XLSX"""
    rich.print("Crear un archivo XLSX")

    # Obtener configuracion
    settings = get_settings()

    # Buscar RFC
    try:
        dispersiones = buscar_rfc(settings=settings, rfc=rfc)
    except MyAnyError as error:
        typer.secho(str(error), fg=typer.colors.RED)
        raise typer.Exit()

    # Crear el libro
    wb = Workbook()

    # Bucle para iterar las dispersiones
    hoja_contador = 1
    for dispersion in dispersiones:
        # Crear una hoja nueva
        ws = wb.create_sheet(title=f"Hoja {hoja_contador}")

        # Tomar el RFC de la persona, va a ser el nombre del archivo que se guardara
        rfc = dispersion.persona.rfc

        # Escribir los datos de la persona
        ws.append(["Centro de trabajo", dispersion.persona.centro_trabajo_clave])
        ws.append(["RFC", rfc])
        ws.append(["Nombre", dispersion.persona.nombre])
        ws.append(["Municipio", dispersion.persona.municipio.nombre])
        ws.append(["Plaza", dispersion.persona.plaza])
        ws.append(["Sexo", dispersion.persona.sexo])

        # Escribir las percepciones
        ws.append(["Percepciones"])
        for percepcion in dispersion.percepciones_deducciones:
            if percepcion.concepto.p_d.value == "P":
                ws.append([percepcion.concepto.descripcion, percepcion.importe])

        # Escribir las deducciones
        ws.append(["Deducciones"])
        for percepcion in dispersion.percepciones_deducciones:
            if percepcion.concepto.p_d.value == "D":
                ws.append([percepcion.concepto.descripcion, percepcion.importe])

        # Escribir las cantidades finales
        ws.append(["Percepcion", dispersion.percepcion])
        ws.append(["Deduccion", dispersion.deduccion])
        ws.append(["Importe", dispersion.importe])
        ws.append(["No. Cheque", dispersion.num_cheque])

        # Incrementar el contador de hojas
        hoja_contador += 1

    # Guardar el archivo, con el nombre del archivo con el RFC
    wb.save(f"{rfc}.xlsx")


@app.command()
def generar():
    """Crear layout para el banco"""
    rich.print("Crear layout para el banco")

    # Obtener configuracion
    settings = get_settings()

    # Leer todo el archivo de explotacion
    try:
        dispersiones = leer(settings=settings)
    except MyAnyError as error:
        typer.secho(str(error), fg=typer.colors.RED)
        raise typer.Exit()

    # Crear el libro
    wb = Workbook()
    ws = wb.active

    # Escribir los encabezados en el primer renglon
    ws.append(
        [
            "Centro de trabajo",
            "RFC",
            "Nombre",
            "Municipio",
            "Plaza",
            "Sexo",
            "Percepcion",
            "Deduccion",
            "Importe",
            "No. Cheque",
        ]
    )

    # Bucle para iterar las dispersiones
    for dispersion in dispersiones:
        # Si uno de los conceptos es "ME", se salta la dispersion
        tiene_monedero_electronico = False
        for percepcion_deduccion in dispersion.percepciones_deducciones:
            if percepcion_deduccion.concepto.concepto == "ME":
                tiene_monedero_electronico = True
                break
        if tiene_monedero_electronico:
            continue

        # Escribir los datos de la persona
        ws.append(
            [
                dispersion.persona.centro_trabajo_clave,
                dispersion.persona.rfc,
                dispersion.persona.nombre,
                dispersion.persona.municipio.nombre,
                dispersion.persona.plaza,
                dispersion.persona.sexo,
                dispersion.percepcion,
                dispersion.deduccion,
                dispersion.importe,
                dispersion.num_cheque,
            ]
        )

    # Guardar el archivo, con el nombre TODOS.xlsx
    wb.save("bancos.xlsx")
